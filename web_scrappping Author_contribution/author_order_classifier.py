Classifies whether a journal uses ALPHABETICAL/RANDOM (A/R), RELATIVE CONTRIBUTION (RC),
or EXPLICIT CONTRIBUTION (EC) author ordering by analysing a sample of papers.

5-Step Per-Paper Classification Pipeline:
  Step 1  — Fetch paper metadata from Crossref / Semantic Scholar / Google Scholar
  Step 2  — Filter: only papers with >= 4 authors
  Step 3  — Fetch full text via Unpaywall (open-access) → fallback to DOI landing page
             Search for >= 2 CRediT keywords  →  EC-CRediT ✅
  Step 4  — If no CRediT: search for ACI markers (phrase + %)  →  EC-ACI ✅
  Step 5  — Check alphabetical name ordering  →  A/R ✅
  Step 6  — Default  →  RC ✅

Journal-Level Decision (applied after classifying all papers):
  EC_rate  >= 30%  →  Explicit Contribution (EC)
  AR_rate  >= 75%  →  Alphabetical / Random (A/R)
  AR_rate  <= 25%  →  Relative Contribution (RC)
  Otherwise        →  Mixed / Hybrid

False-Positive Probability (random list appearing alphabetical by chance):
  n=4  →  1/24  ≈  4.2%   (minimum threshold)
  n=5  →  1/120 ≈  0.8%
  n=6+ →  < 0.1%

Unpaywall:
  Free, no API key — just provide your email in .env as contact_email.
  Covers PubMed Central, arXiv, institutional repositories, author pages.
  ~50% of recent papers have an open-access version somewhere.

Usage:
  python author_order_classifier.py --issn 1932-6203 --max 200 --sample 80
  python author_order_classifier.py --journal "Nature Communications" --max 300 --sample 100
  python author_order_classifier.py --journal "Frontiers in Medicine" --max 200 --sample 50
  python author_order_classifier.py   # interactive prompt
"""

import re
import math
import argparse
import os
import time
import random
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv

try:
    from scholarly import scholarly
    SCHOLARLY_AVAILABLE = True
except ImportError:
    SCHOLARLY_AVAILABLE = False
    print("⚠️  scholarly not installed. Google Scholar support disabled.")
    print("   Install with: pip install scholarly")

load_dotenv()

CROSSREF_BASE         = "https://api.crossref.org/works"
SEMANTIC_SCHOLAR_BASE = "https://api.semanticscholar.org/graph/v1"
UNPAYWALL_BASE        = "https://api.unpaywall.org/v2"
CONTACT_EMAIL         = os.getenv("contact_email", "researcher@example.com")


# ── All 14 CRediT Roles + Section Header Phrases ─────────────────────────────

CREDIT_KEYWORDS = [
    # 14 standardised CRediT roles
    "conceptualization",
    "data curation",
    "formal analysis",
    "funding acquisition",
    "investigation",
    "methodology",
    "project administration",
    "resources",
    "software",
    "supervision",
    "validation",
    "visualization",
    "writing – original draft",
    "writing - original draft",
    "writing – review",
    "writing - review",
    "review & editing",
    "review and editing",
    # Section header phrases that signal a CRediT block (1 match is sufficient)
    "author contributions",
    "authors' contributions",
    "contributions of authors",
]

# Section headers that alone are enough to confirm a CRediT block
CREDIT_SECTION_HEADERS = {
    "author contributions",
    "authors' contributions",
    "contributions of authors",
}


# ── ACI Detection ─────────────────────────────────────────────────────────────
# Requires BOTH a phrase signal AND a percentage value to avoid false positives
# from papers that merely mention percentages in their results sections.

ACI_PHRASES = [
    "author contribution index",
    "percentage contribution",
    "percentage of contribution",
    "contribution percentage",
    "authors' percentage",
    "author percentage",
    "% of the work",
    "% of this work",
    "% contribution",
]

# Matches numbers like "50%", "30 %", "12.5%"
ACI_PERCENT_PATTERN = re.compile(r'\b\d{1,3}(\.\d+)?\s*%')


def check_aci_markers(html_text: str) -> bool:
    """
    Return True only when BOTH signals are present:
      1. At least one ACI phrase keyword
      2. At least one percentage value in the same document

    Requiring both signals prevents false positives from papers that
    mention percentages in figures, statistics, or results sections.
    """
    matched_phrases = [ph for ph in ACI_PHRASES if ph in html_text]
    has_percentages = bool(ACI_PERCENT_PATTERN.search(html_text))
    return len(matched_phrases) >= 1 and has_percentages


# ── Step 3a: Unpaywall — Free Open-Access Full-Text Resolver ─────────────────

def get_fulltext_url_unpaywall(doi: str) -> str | None:
    """
    Query the Unpaywall API to find a legal open-access version of the paper.

    Returns the best OA PDF/HTML URL if found, or None if not available.
    Covers: PubMed Central, arXiv, institutional repositories, author pages.
    Free to use — no API key needed, just a contact email.
    """
    if not doi or doi == "N/A":
        return None
    try:
        url = f"{UNPAYWALL_BASE}/{doi}"
        resp = requests.get(
            url,
            params={"email": CONTACT_EMAIL},
            timeout=10,
        )
        if resp.status_code != 200:
            return None
        data = resp.json()
        # is_oa=True means at least one open-access location exists
        if not data.get("is_oa"):
            return None
        best = data.get("best_oa_location") or {}
        return best.get("url_for_landing_page") or best.get("url") or None
    except Exception:
        return None


# ── Step 3b: Fetch Full Text (Unpaywall → DOI Landing Page Fallback) ──────────

def fetch_fulltext(doi: str) -> tuple[str, str]:
    """
    Attempt to fetch the paper's full text.

    Strategy:
      1. Query Unpaywall for an open-access URL
      2. If found, fetch that URL (full text likely available)
      3. If not found, fall back to the DOI landing page (abstract only ⚠️)

    Returns (html_lower, source) where source is "unpaywall", "doi", or "failed".
    """
    if not doi or doi == "N/A":
        return "", "failed"

    headers = {
        "User-Agent": (
            f"AuthorOrderClassifier/2.0 "
            f"(mailto:{CONTACT_EMAIL}; academic research bot)"
        ),
        "Accept": "text/html,application/xhtml+xml",
    }

    # Try Unpaywall first
    oa_url = get_fulltext_url_unpaywall(doi)
    if oa_url:
        try:
            resp = requests.get(oa_url, headers=headers, timeout=15, allow_redirects=True)
            if resp.status_code == 200:
                return resp.text.lower(), "unpaywall"
        except Exception:
            pass

    # Fallback: DOI landing page
    try:
        resp = requests.get(
            f"https://doi.org/{doi}",
            headers=headers,
            timeout=15,
            allow_redirects=True,
        )
        if resp.status_code == 200:
            return resp.text.lower(), "doi"
    except Exception:
        pass

    return "", "failed"


# ── Step 3c: CRediT Keyword Scan ──────────────────────────────────────────────

def check_credit_keywords(html_lower: str) -> bool:
    """
    Return True if the full text contains evidence of a CRediT section.

    Two-tier check:
      1. If ANY section-header phrase is found ("author contributions", etc.)
         → immediately True (these headers only appear in contribution sections).
      2. Otherwise require >= 2 CRediT role keywords (e.g. "conceptualization"
         + "supervision") to confirm a role-based contribution block.

    This avoids false positives from papers that mention a single role word
    (e.g. "software" or "resources") in an unrelated context.
    """
    # Tier 1: section header alone is conclusive
    for header in CREDIT_SECTION_HEADERS:
        if header in html_lower:
            return True

    # Tier 2: need >= 2 CRediT role keywords
    role_keywords = [kw for kw in CREDIT_KEYWORDS if kw not in CREDIT_SECTION_HEADERS]
    matched_roles = [kw for kw in role_keywords if kw in html_lower]
    return len(matched_roles) >= 2


# ── Steps 5 & 6: Author Name Normalisation & Alphabetical Check ───────────────

def normalize_family_name(name: str) -> str:
    """Lowercase, remove punctuation and special characters, strip whitespace."""
    name = name.lower()
    name = re.sub(r"[^a-z\s]", "", name)
    return name.strip()


def is_alphabetical(authors: list) -> bool:
    """
    Return True if authors are sorted A→Z by normalised family name.

    Guards:
      - Requires >= 4 authors (fewer authors make the test statistically
        unreliable — 3 authors have a 1/6 = 16.7% false-positive rate).
      - Returns False if any author is missing a family name.
    """
    if len(authors) < 4:
        return False  # not enough authors to reliably detect alphabetical ordering

    family_names = []
    for author in authors:
        family = author.get("family", "").strip()
        if not family:
            return False
        family_names.append(normalize_family_name(family))
    return family_names == sorted(family_names)


def chance_probability(n: int) -> float:
    """Probability that n randomly ordered authors appear alphabetical = 1/n!"""
    if n > 20:
        return 1e-20
    try:
        return 1.0 / math.factorial(n)
    except (OverflowError, ValueError):
        return 1e-20


# ── Per-Paper Classification ───────────────────────────────────────────────────

def classify_paper(authors: list, has_credit: bool, has_aci: bool) -> dict:
    """
    5-step per-paper classification:

      Step 3 — EC-CRediT : >= 2 CRediT keywords found in full text
      Step 4 — EC-ACI    : ACI phrase + percentage markers found (only if no CRediT)
      Step 5 — A/R       : authors sorted A→Z (no EC signal)
      Step 6 — RC        : default (not alphabetical, no EC signal)

    Papers not in the full-text sample have has_credit=None, has_aci=None and
    go directly to Steps 5→6 (alphabetical check then RC default).
    """
    n = len(authors)

    # Step 3: EC-CRediT
    if has_credit is True:
        return {
            "classification": "EC",
            "ec_type":        "CRediT",
            "ordering_type":  "Explicit Contribution (EC-CRediT)",
            "is_alphabetical": is_alphabetical(authors),
            "confidence":     0.95,
            "chance_prob":    round(chance_probability(n), 8),
            "interpretation": (
                ">= 2 CRediT taxonomy keywords found in full text → "
                "Explicit Contribution (EC-CRediT)."
            ),
        }

    # Step 4: EC-ACI (only checked when CRediT not found)
    if has_aci is True:
        return {
            "classification": "EC",
            "ec_type":        "ACI",
            "ordering_type":  "Explicit Contribution (EC-ACI)",
            "is_alphabetical": is_alphabetical(authors),
            "confidence":     0.90,
            "chance_prob":    round(chance_probability(n), 8),
            "interpretation": (
                "ACI phrase + percentage markers found → "
                "Explicit Contribution (EC-ACI)."
            ),
        }

    # Step 5: Alphabetical / Random
    is_alpha = is_alphabetical(authors)
    fp_prob  = chance_probability(n)
    if is_alpha:
        confidence = min(round(1.0 - fp_prob, 4), 0.999)
        return {
            "classification": "A/R",
            "ec_type":        None,
            "ordering_type":  "Alphabetical / Random (A/R)",
            "is_alphabetical": True,
            "confidence":     confidence,
            "chance_prob":    round(fp_prob, 8),
            "interpretation": (
                f"Authors sorted A→Z. "
                f"False-positive prob = 1/{n}! = {fp_prob:.2e}."
            ),
        }

    # Step 6: Relative Contribution (default)
    confidence = 0.90 if n >= 6 else 0.75
    return {
        "classification": "RC",
        "ec_type":        None,
        "ordering_type":  "Relative Contribution (RC)",
        "is_alphabetical": False,
        "confidence":     confidence,
        "chance_prob":    round(fp_prob if is_alpha else chance_probability(n), 8),
        "interpretation": (
            "Not alphabetical, no CRediT/ACI signal → "
            "Relative Contribution (RC). "
            "1st author = main contributor; last author = senior supervisor."
        ),
    }


# ── Step 1a: Fetch from Crossref ──────────────────────────────────────────────

MIN_ELIGIBLE_PAPERS = 20   # must reach this many 4+-author papers before stopping


def fetch_papers_crossref(issn=None, journal_name=None, max_papers=500,
                          min_eligible=MIN_ELIGIBLE_PAPERS, min_authors=4):
    """
    Fetch paper metadata from the Crossref API.

    Keeps fetching beyond `max_papers` if needed until at least `min_eligible`
    papers with >= `min_authors` authors have been collected, so that journals
    full of single-author editorials/letters never produce "Insufficient data".
    A hard cap of max(max_papers * 5, 1500) prevents runaway fetching.
    """
    papers = []
    eligible_count = 0
    rows_per_page  = 100
    offset         = 0
    hard_cap       = max(max_papers * 5, 1500)

    filters = ["type:journal-article"]
    if issn:
        filters.append(f"issn:{issn}")
    filter_str = ",".join(filters)

    base_params = {
        "filter":  filter_str,
        "rows":    rows_per_page,
        "select":  "title,author,published-print,published-online,DOI,container-title",
        "mailto":  CONTACT_EMAIL,
    }
    if journal_name and not issn:
        base_params["query.container-title"] = journal_name

    print(f"\n📡 [Crossref] Fetching papers (target: {max_papers}, need >= {min_eligible} eligible)...")

    while len(papers) < hard_cap:
        params = {**base_params, "offset": offset}
        try:
            resp = requests.get(CROSSREF_BASE, params=params, timeout=20)
            if resp.status_code != 200:
                print(f"    Crossref HTTP {resp.status_code}, stopping.")
                break
            data  = resp.json()
            items = data.get("message", {}).get("items", [])
            if not items:
                break

            for item in items:
                authors = item.get("author", [])
                if not authors:
                    continue
                pub_date   = item.get("published-print") or item.get("published-online") or {}
                year_parts = pub_date.get("date-parts", [[None]])
                year       = year_parts[0][0] if year_parts and year_parts[0] else None
                container  = item.get("container-title", [])
                journal    = container[0] if container else "N/A"
                papers.append({
                    "doi":     item.get("DOI", "N/A"),
                    "title":   (item.get("title") or ["N/A"])[0],
                    "authors": authors,
                    "year":    year,
                    "journal": journal,
                    "source":  "Crossref",
                })
                if len(authors) >= min_authors:
                    eligible_count += 1

            offset += rows_per_page
            print(f"  {len(papers)} papers fetched ({eligible_count} eligible)...")

            # Stop once we have enough total AND enough eligible
            if len(papers) >= max_papers and eligible_count >= min_eligible:
                break

            if len(items) < rows_per_page:
                break

        except requests.exceptions.Timeout:
            print("  ⚠️  Crossref timeout. Stopping.")
            break
        except Exception as e:
            print(f"  ⚠️  Crossref error: {e}")
            break

    if eligible_count < min_eligible:
        print(f"  ⚠️  Only {eligible_count} eligible papers found (need {min_eligible}). "
              f"Journal may have few multi-author papers in Crossref.")

    print(f"✅ Crossref: {len(papers)} papers ({eligible_count} with >= {min_authors} authors).")
    return papers


# ── Step 1b: Fetch from Semantic Scholar ─────────────────────────────────────

def fetch_papers_semantic_scholar(journal_name: str, max_papers: int = 200) -> list:
    """Fetch paper metadata from the Semantic Scholar API."""
    papers = []
    fields = "title,authors,year,externalIds,venue"
    limit  = min(100, max_papers)
    offset = 0

    print(f"\n📡 [Semantic Scholar] Fetching papers (target: {max_papers})...")

    while len(papers) < max_papers:
        try:
            params = {"query": journal_name, "fields": fields,
                      "limit": limit, "offset": offset}
            resp = requests.get(
                f"{SEMANTIC_SCHOLAR_BASE}/paper/search",
                params=params, timeout=20,
                headers={"User-Agent": f"AuthorOrderClassifier/2.0 (mailto:{CONTACT_EMAIL})"},
            )
            if resp.status_code == 429:
                wait = int(resp.headers.get("Retry-After", 15))
                print(f"  Rate-limited. Waiting {wait}s...")
                time.sleep(wait)
                continue
            if resp.status_code != 200:
                print(f"    Semantic Scholar HTTP {resp.status_code}, stopping.")
                break

            items = resp.json().get("data", [])
            if not items:
                break

            for item in items:
                raw_authors = item.get("authors", [])
                if not raw_authors:
                    continue
                authors = []
                for a in raw_authors:
                    parts = a.get("name", "").strip().split()
                    if parts:
                        authors.append({
                            "family": parts[-1],
                            "given":  " ".join(parts[:-1]) if len(parts) > 1 else "",
                        })
                ext_ids = item.get("externalIds") or {}
                papers.append({
                    "doi":     ext_ids.get("DOI", "N/A"),
                    "title":   item.get("title", "N/A") or "N/A",
                    "authors": authors,
                    "year":    item.get("year"),
                    "journal": item.get("venue", "N/A") or "N/A",
                    "source":  "SemanticScholar",
                })

            offset += len(items)
            print(f"  {len(papers)} papers fetched...")
            if len(items) < limit:
                break
            time.sleep(1)

        except requests.exceptions.Timeout:
            print("    Semantic Scholar timeout.")
            break
        except Exception as e:
            print(f"    Semantic Scholar error: {e}")
            break

    print(f"✅ Semantic Scholar: {len(papers)} papers.")
    return papers[:max_papers]


# ── Step 1c: Fetch from Google Scholar ───────────────────────────────────────

def fetch_papers_google_scholar(journal_name: str, max_papers: int = 100) -> list:
    """Fetch paper metadata from Google Scholar via the scholarly library."""
    if not SCHOLARLY_AVAILABLE:
        print("⚠️  scholarly not available. Skipping Google Scholar.")
        return []

    papers = []
    print(f"\n📡 [Google Scholar] Fetching papers (target: {max_papers})...")

    try:
        search_query = scholarly.search_pubs(f'"{journal_name}"')
        for pub in search_query:
            if len(papers) >= max_papers:
                break
            try:
                authors = []
                raw_authors = pub.get("author", [])
                if isinstance(raw_authors, str):
                    for name in raw_authors.split(" and "):
                        name = name.strip()
                        if name:
                            parts  = name.split()
                            family = parts[-1]
                            given  = " ".join(parts[:-1]) if len(parts) > 1 else ""
                            authors.append({"family": family, "given": given})
                elif isinstance(raw_authors, list):
                    for author in raw_authors:
                        if isinstance(author, str):
                            parts  = author.strip().split()
                            if parts:
                                authors.append({
                                    "family": parts[-1],
                                    "given":  " ".join(parts[:-1]) if len(parts) > 1 else "",
                                })
                if not authors:
                    continue
                papers.append({
                    "doi":     pub.get("eprint", "N/A") or "N/A",
                    "title":   pub.get("title",   "N/A") or "N/A",
                    "authors": authors,
                    "year":    pub.get("pub_year"),
                    "journal": pub.get("journal", "N/A") or "N/A",
                    "source":  "GoogleScholar",
                })
                print(f"  {len(papers)} papers fetched...")
                time.sleep(random.uniform(2, 5))
            except Exception:
                continue
    except Exception as e:
        print(f"  ⚠️  Google Scholar error: {e}")
        return []

    print(f"✅ Google Scholar: {len(papers)} papers.")
    return papers[:max_papers]


# ── Steps 2–6: Analyse All Papers ────────────────────────────────────────────

def analyze_papers(papers: list, min_authors: int = 4, sample: int = 50):
    """
    Classify each eligible paper and compute journal-level statistics.

    Full-text checking (Steps 3 & 4) is always enabled.
    The `sample` parameter controls how many papers get the full-text check.
    Papers outside the sample go directly to the alphabetical check (Step 5).

    Journal-Level Decision:
      EC_rate  >= 30%  →  Explicit Contribution (EC)
      AR_rate  >= 75%  →  Alphabetical / Random (A/R)
      AR_rate  <= 25%  →  Relative Contribution (RC)
      Otherwise        →  Mixed / Hybrid

    Paywall Warning:
      If the final classification is RC but >= 50% of sampled papers had
      no open-access full text, a warning is raised — EC papers may have
      been misclassified as RC due to paywall blocking.
    """
    paper_results  = []
    total_eligible = 0
    ec_count        = 0
    ec_credit_count = 0
    ec_aci_count    = 0
    ar_count        = 0
    rc_count        = 0

    # Full-text check tracking
    ft_checked       = 0
    ft_unpaywall_ok  = 0   # fetched via Unpaywall
    ft_doi_ok        = 0   # fetched via DOI landing page only
    ft_failed        = 0   # no text at all (paywall / error)
    credit_hits      = 0
    aci_hits         = 0

    # Select which DOIs get full-text checking (all papers regardless of author count)
    eligible_dois = [
        p["doi"] for p in papers
        if p.get("doi", "N/A") != "N/A"
    ]
    sample_size   = min(sample, len(eligible_dois))
    sampled_dois  = set(random.sample(eligible_dois, sample_size)) if eligible_dois else set()

    print(f"\n  Full-text check (Unpaywall + CRediT + ACI): sampling {len(sampled_dois)} papers...")

    for paper in papers:
        authors = paper.get("authors", [])
        n       = len(authors)

        total_eligible += 1

        has_credit  = None
        has_aci     = None
        ft_source   = "not_checked"

        doi = paper.get("doi", "N/A")
        if doi in sampled_dois:
            ft_checked += 1
            html_lower, ft_source = fetch_fulltext(doi)

            if ft_source == "unpaywall":
                ft_unpaywall_ok += 1
            elif ft_source == "doi":
                ft_doi_ok += 1
            else:
                ft_failed += 1

            if html_lower:
                # Step 3: CRediT
                has_credit = check_credit_keywords(html_lower)
                if has_credit:
                    credit_hits += 1
                else:
                    # Step 4: ACI (only when CRediT not found)
                    has_aci = check_aci_markers(html_lower)
                    if has_aci:
                        aci_hits += 1

            signal = (
                f"EC-CRediT [{ft_source}]" if has_credit
                else f"EC-ACI [{ft_source}]" if has_aci
                else f"No EC [{ft_source}]"
            )
            print(f"  [{ft_checked}/{len(sampled_dois)}] {doi[:40]}... → {signal}")
            time.sleep(random.uniform(1.5, 3.0))

        clf = classify_paper(authors, has_credit, has_aci)

        if clf["classification"] == "EC":
            ec_count += 1
            if clf["ec_type"] == "CRediT":
                ec_credit_count += 1
            elif clf["ec_type"] == "ACI":
                ec_aci_count += 1
        elif clf["classification"] == "A/R":
            ar_count += 1
        elif clf["classification"] == "RC":
            rc_count += 1

        paper_results.append({
            "doi":            doi,
            "title":          paper.get("title",   "N/A"),
            "journal":        paper.get("journal", "N/A"),
            "year":           paper.get("year"),
            "num_authors":    n,
            "author_names":   " | ".join(a.get("family", "?") for a in authors),
            "is_alphabetical": clf["is_alphabetical"],
            "classification": clf["classification"],
            "ec_type":        clf.get("ec_type") or "—",
            "ordering_type":  clf["ordering_type"],
            "confidence":     clf["confidence"],
            "chance_prob":    clf["chance_prob"],
            "fulltext_source": ft_source,
            "source":         paper.get("source", "N/A"),
            "interpretation": clf["interpretation"],
        })

    # ── Rates ────────────────────────────────────────────────────────────────
    def rate(count):
        return round(count / total_eligible, 4) if total_eligible > 0 else 0.0

    ec_rate = rate(ec_count)
    ar_rate = rate(ar_count)
    rc_rate = rate(rc_count)

    paywall_rate = round(ft_failed / ft_checked, 4) if ft_checked > 0 else None

    # ── Journal-Level Decision ────────────────────────────────────────────────
    if total_eligible < 20:
        conclusion = "Insufficient data"
        confidence = "Low"
    elif ec_rate >= 0.30:
        conclusion = "Explicit Contribution (EC)"
        confidence = "High" if total_eligible >= 100 else "Medium"
    elif ar_rate >= 0.75:
        conclusion = "Alphabetical / Random (A/R)"
        confidence = "High" if total_eligible >= 100 else "Medium"
    elif ar_rate <= 0.25:
        conclusion = "Relative Contribution (RC)"
        confidence = "High" if total_eligible >= 100 else "Medium"
    else:
        conclusion = "Mixed / Hybrid"
        confidence = "Low"

    # ── Paywall Warning ───────────────────────────────────────────────────────
    paywall_warning = (
        "RC" in conclusion
        and paywall_rate is not None
        and paywall_rate >= 0.50
    )

    summary = {
        "total_fetched":      len(papers),
        "eligible_papers":    total_eligible,
        # Paper counts
        "ec_papers":          ec_count,
        "ec_credit_papers":   ec_credit_count,
        "ec_aci_papers":      ec_aci_count,
        "ar_papers":          ar_count,
        "rc_papers":          rc_count,
        # Rates
        "ec_rate":            ec_rate,
        "ar_rate":            ar_rate,
        "rc_rate":            rc_rate,
        # Full-text stats
        "ft_checked":         ft_checked,
        "ft_unpaywall_ok":    ft_unpaywall_ok,
        "ft_doi_ok":          ft_doi_ok,
        "ft_failed":          ft_failed,
        "paywall_rate":       paywall_rate,
        "credit_hits":        credit_hits,
        "aci_hits":           aci_hits,
        # Decision
        "conclusion":         conclusion,
        "confidence":         confidence,
        "min_authors_filter": min_authors,
        "paywall_warning":    paywall_warning,
    }

    return paper_results, summary


# ── Export to Excel ───────────────────────────────────────────────────────────

def save_to_excel(paper_results: list, summary: dict, journal_label: str):
    """
    Save results to a styled Excel workbook with two sheets:
      Sheet 1 — Journal Summary
      Sheet 2 — Per-Paper Results
    """
    wb = Workbook()

    # Colour palette
    dark_blue    = "1F4E79"
    mid_blue     = "366092"
    green_fill   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    blue_fill    = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    orange_fill  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    ec_fill      = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    aci_fill     = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    ar_fill      = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
    rc_fill      = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_align    = Alignment(vertical="top", wrap_text=True)

    def header_style(ws, row_num, bg_color):
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=12)
        for cell in ws[row_num]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = center_align

    def clf_fill(classification):
        if classification == "EC":    return ec_fill
        if classification == "A/R":   return ar_fill
        if classification == "RC":    return rc_fill
        return yellow_fill

    conclusion = summary["conclusion"]
    result_fill = (
        green_fill  if "EC"          in conclusion else
        blue_fill   if "A/R"         in conclusion else
        yellow_fill if "RC"          in conclusion else
        red_fill
    )

    # ── Sheet 1: Journal Summary ──────────────────────────────────────────────
    ws1       = wb.active
    ws1.title = "Journal Summary"
    ws1.append(["Metric", "Value"])
    header_style(ws1, 1, dark_blue)

    ws1.append(["── 5-Step Author Ordering Analysis ──", ""])
    header_style(ws1, ws1.max_row, dark_blue)

    stats_rows = [
        ("Journal / Query",                              journal_label),
        ("Min Authors Filter",                           f">= {summary['min_authors_filter']} authors"),
        ("Total Papers Fetched",                         summary["total_fetched"]),
        ("Eligible Papers",                              summary["eligible_papers"]),
        ("", ""),
        ("EC  Papers — total (Explicit Contribution)",  summary["ec_papers"]),
        ("  ↳ EC-CRediT  (Step 3 — CRediT keywords)",  summary["ec_credit_papers"]),
        ("  ↳ EC-ACI     (Step 4 — ACI % markers)",    summary["ec_aci_papers"]),
        ("A/R Papers     (Step 5 — Alphabetical)",      summary["ar_papers"]),
        ("RC  Papers     (Step 6 — Default RC)",        summary["rc_papers"]),
        ("", ""),
        ("EC  Rate  (threshold >= 30%)",                f"{summary['ec_rate']:.2%}"),
        ("A/R Rate  (threshold >= 75%)",                f"{summary['ar_rate']:.2%}"),
        ("RC  Rate  (threshold <= 25%)",                f"{summary['rc_rate']:.2%}"),
        ("", ""),
        ("JOURNAL CLASSIFICATION",                      conclusion),
        ("Confidence",                                   summary["confidence"]),
    ]

    fill_map = {
        "EC  Rate  (threshold >= 30%)": ec_fill,
        "A/R Rate  (threshold >= 75%)": ar_fill,
        "RC  Rate  (threshold <= 25%)": rc_fill,
    }

    for metric, value in stats_rows:
        ws1.append([metric, value])
        row_idx = ws1.max_row
        if metric in ("JOURNAL CLASSIFICATION", "Confidence"):
            for col in [1, 2]:
                ws1.cell(row_idx, col).fill = result_fill
                ws1.cell(row_idx, col).font = Font(bold=True, size=12)
        elif metric in fill_map:
            for col in [1, 2]:
                ws1.cell(row_idx, col).fill = fill_map[metric]
        for col in [1, 2]:
            ws1.cell(row_idx, col).alignment = center_align

    # ── Paywall Warning ──
    ws1.append([])
    ws1.append(["── Unpaywall & Full-text Access ──", ""])
    header_style(ws1, ws1.max_row, "ED7D31")

    paywall_rate    = summary.get("paywall_rate")
    pw_rate_display = f"{paywall_rate:.2%}" if paywall_rate is not None else "N/A"

    ft_rows = [
        ("Papers Sampled for Full-text Check",   summary["ft_checked"]),
        ("Fetched via Unpaywall (open-access)",   summary["ft_unpaywall_ok"]),
        ("Fetched via DOI landing page only",     summary["ft_doi_ok"]),
        ("Failed / Paywalled (no full text)",     summary["ft_failed"]),
        ("Paywall Block Rate",                    pw_rate_display),
        ("Papers with CRediT Section Found",      summary["credit_hits"]),
        ("Papers with ACI Markers Found",         summary["aci_hits"]),
    ]
    for metric, value in ft_rows:
        ws1.append([metric, value])
        row_idx = ws1.max_row
        if metric == "Failed / Paywalled (no full text)" and summary["ft_failed"] > 0:
            for col in [1, 2]:
                ws1.cell(row_idx, col).fill = orange_fill
        for col in [1, 2]:
            ws1.cell(row_idx, col).alignment = center_align

    if summary.get("paywall_warning"):
        ws1.append([])
        ws1.append([
            "⚠️  PAYWALL WARNING",
            (
                "More than 50% of sampled papers had no open-access full text. "
                "EC papers may have been misclassified as RC. "
                "Try a fully open-access journal (e.g. PLOS ONE, PeerJ, Frontiers)."
            ),
        ])
        row_idx = ws1.max_row
        for col in [1, 2]:
            ws1.cell(row_idx, col).fill = red_fill
            ws1.cell(row_idx, col).font = Font(bold=True, size=11, color="9C0006")
            ws1.cell(row_idx, col).alignment = center_align

    # ── Classification Legend ──
    ws1.append([])
    ws1.append(["Step", "Condition", "Classification", "Notes"])
    header_style(ws1, ws1.max_row, mid_blue)

    legend = [
        ("Step 3",        ">= 2 CRediT keywords in full text",    "EC-CRediT",           "Highest confidence (0.95)"),
        ("Step 4",        "ACI phrase + % markers in full text",   "EC-ACI",              "High confidence (0.90)"),
        ("Step 5",        "Authors sorted A→Z (no EC signal)",     "Alphabetical / Random","Confidence = 1 - 1/n!"),
        ("Step 6",        "Default (not alphabetical, no EC)",     "Relative Contribution","n>=6: 0.90, else 0.75"),
        ("Journal: EC",   "EC_rate  >= 30%",                       "Explicit Contribution",""),
        ("Journal: A/R",  "A/R_rate >= 75%",                       "Alphabetical / Random",""),
        ("Journal: RC",   "A/R_rate <= 25%",                       "Relative Contribution",""),
        ("Journal: Mixed","26%–74% A/R, EC < 30%",                "Mixed / Hybrid",       ""),
    ]
    legend_fills = [ec_fill, aci_fill, ar_fill, rc_fill, ec_fill, ar_fill, rc_fill, yellow_fill]
    for row_data, lfill in zip(legend, legend_fills):
        ws1.append(list(row_data))
        row_idx = ws1.max_row
        for col in range(1, 5):
            ws1.cell(row_idx, col).fill      = lfill
            ws1.cell(row_idx, col).alignment = center_align

    ws1.column_dimensions["A"].width = 48
    ws1.column_dimensions["B"].width = 42
    ws1.column_dimensions["C"].width = 26
    ws1.column_dimensions["D"].width = 32

    # ── Sheet 2: Per-Paper Results ────────────────────────────────────────────
    ws2       = wb.create_sheet("Per-Paper Results")
    headers   = [
        "No.",                            # 1
        "DOI",                            # 2
        "Publication Title",              # 3
        "Journal",                        # 4
        "Year",                           # 5
        "# Authors",                      # 6
        "Author Family Names (in order)", # 7
        "Is Alphabetical?",               # 8
        "Classification",                 # 9
        "EC Type (CRediT / ACI)",         # 10
        "Ordering Type",                  # 11
        "Confidence Score",               # 12
        "Chance Prob (1/n!)",             # 13
        "Full-text Source",               # 14
        "Data Source",                    # 15
        "Interpretation",                 # 16
    ]
    ws2.append(headers)
    header_style(ws2, 1, mid_blue)

    for i, p in enumerate(paper_results, 1):
        classification = p.get("classification", "—")
        ec_type        = p.get("ec_type", "—")

        ws2.append([
            i,
            p["doi"],
            p["title"],
            p["journal"],
            p["year"],
            p["num_authors"],
            p["author_names"],
            "Yes" if p["is_alphabetical"] else "No",
            classification,
            ec_type,
            p.get("ordering_type", "—"),
            p.get("confidence", 0.0),
            p.get("chance_prob", 0.0),
            p.get("fulltext_source", "not_checked"),
            p.get("source", "N/A"),
            p.get("interpretation", ""),
        ])
        row_idx = ws2.max_row

        # Is Alphabetical? (col 8)
        cell8 = ws2.cell(row_idx, 8)
        cell8.fill      = blue_fill if p["is_alphabetical"] else red_fill
        cell8.font      = Font(bold=True)
        cell8.alignment = center_align

        # Classification (col 9)
        cell9 = ws2.cell(row_idx, 9)
        cell9.fill      = clf_fill(classification)
        cell9.font      = Font(bold=True)
        cell9.alignment = center_align

        # EC Type (col 10)
        cell10 = ws2.cell(row_idx, 10)
        if ec_type == "CRediT":
            cell10.fill = ec_fill
            cell10.font = Font(bold=True)
        elif ec_type == "ACI":
            cell10.fill = aci_fill
            cell10.font = Font(bold=True)
        cell10.alignment = center_align

        # Full-text source (col 14) — colour by quality
        cell14 = ws2.cell(row_idx, 14)
        if p.get("fulltext_source") == "unpaywall":
            cell14.fill = green_fill
        elif p.get("fulltext_source") == "doi":
            cell14.fill = yellow_fill
        elif p.get("fulltext_source") == "failed":
            cell14.fill = red_fill
        cell14.alignment = center_align

        for col in range(1, len(headers) + 1):
            ws2.cell(row_idx, col).alignment = (
                top_align if col in [3, 7, 11, 16] else center_align
            )

    col_widths = [5, 28, 52, 28, 7, 9, 55, 15, 20, 16, 34, 14, 16, 16, 14, 55]
    for col_idx, width in enumerate(col_widths, 1):
        ws2.column_dimensions[ws2.cell(1, col_idx).column_letter].width = width

    # ── Save ──
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = re.sub(r"[^\w]", "_", journal_label)
    filename   = f"author_order_{safe_label}_{timestamp}.xlsx"
    wb.save(filename)
    print(f"\n📁 Results saved to: {filename}")
    return filename


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Classify journal author ordering (EC / A/R / RC) via Crossref + Unpaywall."
    )
    parser.add_argument("--issn",    type=str, help="Journal ISSN (e.g. 1932-6203 for PLOS ONE)")
    parser.add_argument("--journal", type=str, help="Journal name keyword")
    parser.add_argument("--max",     type=int, default=500, help="Max papers to fetch (default: 500)")
    parser.add_argument("--min-authors", type=int, default=4,
                        help="Min authors per paper to include (default: 4)")
    parser.add_argument("--source",  choices=["crossref", "semantic", "google", "both", "all"],
                        default="crossref",
                        help="Data source (default: crossref)")
    parser.add_argument("--sample",  type=int, default=50,
                        help="Papers to sample for full-text CRediT+ACI check (default: 50)")
    args = parser.parse_args()

    if not args.issn and not args.journal:
        user_input = input("Enter journal name or ISSN: ").strip()
        if user_input.replace("-", "").isdigit() or (
            len(user_input) == 9 and user_input[4] == "-"
        ):
            args.issn = user_input
        else:
            args.journal = user_input

    journal_label = args.issn or args.journal

    print(f"\n{'='*62}")
    print(f"  Author Order Classifier")
    print(f"  Journal        : {journal_label}")
    print(f"  Max papers     : {args.max}")
    print(f"  Min authors    : {args.min_authors}+")
    print(f"  Source         : {args.source}")
    print(f"  Full-text sample: {args.sample} papers (Unpaywall enabled)")
    print(f"{'='*62}")

    # ── Step 1: Fetch papers ──
    papers = []

    if args.source in ("crossref", "both", "all"):
        papers.extend(fetch_papers_crossref(
            issn=args.issn, journal_name=args.journal, max_papers=args.max,
            min_eligible=MIN_ELIGIBLE_PAPERS, min_authors=args.min_authors,
        ))

    if args.source in ("semantic", "both", "all"):
        sem_max    = args.max if args.source == "semantic" else max(100, args.max // 3)
        sem_papers = fetch_papers_semantic_scholar(journal_label, sem_max)
        existing   = {p["doi"] for p in papers if p["doi"] != "N/A"}
        new_papers = [p for p in sem_papers if p["doi"] not in existing]
        papers.extend(new_papers)
        print(f"  Added {len(new_papers)} unique papers from Semantic Scholar.")

    if args.source in ("google", "all"):
        goog_max    = args.max if args.source == "google" else max(50, args.max // 5)
        goog_papers = fetch_papers_google_scholar(journal_label, goog_max)
        existing    = {p["doi"] for p in papers if p["doi"] != "N/A"}
        new_papers  = [p for p in goog_papers if p["doi"] not in existing]
        papers.extend(new_papers)
        print(f"  Added {len(new_papers)} unique papers from Google Scholar.")

    if not papers:
        print("⚠️  No papers found. Check the ISSN or journal name.")
        return

    # ── Steps 2–6: Classify ──
    paper_results, summary = analyze_papers(
        papers,
        min_authors=args.min_authors,
        sample=args.sample,
    )

    # ── Console Output ──
    conclusion = summary["conclusion"]
    ec_rate    = summary["ec_rate"]
    ar_rate    = summary["ar_rate"]
    rc_rate    = summary["rc_rate"]

    W = 66
    print(f"\n{'='*W}")
    print(f"  5-STEP AUTHOR ORDERING CLASSIFICATION")
    print(f"  Journal : {journal_label}")
    print(f"{'='*W}")
    print(f"  Total papers fetched              : {summary['total_fetched']}")
    print(f"  Eligible papers (>= {args.min_authors} authors)    : {summary['eligible_papers']}")
    print(f"  {'─'*58}")
    print(f"  Per-paper classification:")
    print(f"    EC  total (Explicit Contribution): {summary['ec_papers']:4d}  [{ec_rate:.1%}]")
    print(f"      ↳ EC-CRediT (Step 3)           : {summary['ec_credit_papers']:4d}")
    print(f"      ↳ EC-ACI    (Step 4)           : {summary['ec_aci_papers']:4d}")
    print(f"    A/R (Step 5 — Alphabetical)       : {summary['ar_papers']:4d}  [{ar_rate:.1%}]")
    print(f"    RC  (Step 6 — Default)             : {summary['rc_papers']:4d}  [{rc_rate:.1%}]")
    print(f"  {'─'*58}")
    print(f"  Full-text access (sample={args.sample}):")
    print(f"    Fetched via Unpaywall (OA)       : {summary['ft_unpaywall_ok']}")
    print(f"    Fetched via DOI page (abstract)  : {summary['ft_doi_ok']}")
    print(f"    Failed / Paywalled               : {summary['ft_failed']}")
    if summary.get("paywall_rate") is not None:
        print(f"    Paywall block rate               : {summary['paywall_rate']:.1%}")
    print(f"    CRediT section hits              : {summary['credit_hits']}")
    print(f"    ACI marker hits                  : {summary['aci_hits']}")
    print(f"  {'─'*58}")
    print(f"  Journal-level decision:")
    print(f"    EC_rate  >= 30% : {ec_rate:.1%}  {'<== TRIGGERED' if ec_rate >= 0.30 else ''}")
    print(f"    A/R_rate >= 75% : {ar_rate:.1%}  {'<== TRIGGERED' if ar_rate >= 0.75 else ''}")
    print(f"    A/R_rate <= 25% : {ar_rate:.1%}  {'<== TRIGGERED' if ar_rate <= 0.25 and ec_rate < 0.30 else ''}")
    print(f"  {'─'*58}")
    print(f"  CLASSIFICATION : {conclusion}")
    print(f"  Confidence     : {summary['confidence']}")

    if summary.get("paywall_warning"):
        print(f"\n  ⚠️  WARNING: More than 50% of sampled papers had no open-access full text.")
        print(f"  EC papers may have been misclassified as RC due to paywall blocking.")
        print(f"  Try with a fully open-access journal (e.g. PLOS ONE, PeerJ, Frontiers).")

    print(f"{'='*W}")

    # ── Step 7: Export ──
    save_to_excel(paper_results, summary, journal_label)


if __name__ == "__main__":
    main()
import re
import math
import argparse
import os
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv

load_dotenv()

CROSSREF_BASE = "https://api.crossref.org/works"
CONTACT_EMAIL = os.getenv("contact_email", "researcher@example.com")


# ── Step 1: Fetch metadata from Crossref ────────────────────────────────────

def fetch_papers_crossref(issn=None, journal_name=None, max_papers=1000):
    """
    Fetch paper metadata from the Crossref API.

    Parameters
    ----------
    issn        : str  — Journal ISSN (e.g. "1476-4687")
    journal_name: str  — Journal name keyword (used if no ISSN provided)
    max_papers  : int  — Maximum number of papers to fetch

    Returns
    -------
    list of dicts with keys: doi, title, authors, year
    """
    papers = []
    rows_per_page = 100
    offset = 0

    filters = ["type:journal-article"]
    if issn:
        filters.append(f"issn:{issn}")
    filter_str = ",".join(filters)

    base_params = {
        "filter": filter_str,
        "rows": rows_per_page,
        "select": "title,author,published-print,published-online,DOI,container-title",
        "mailto": CONTACT_EMAIL,  
    }
    if journal_name and not issn:
        base_params["query.container-title"] = journal_name

    print(f"\n📡 Fetching papers from Crossref (target: {max_papers} papers)...")

    while len(papers) < max_papers:
        params = {**base_params, "offset": offset}
        try:
            resp = requests.get(CROSSREF_BASE, params=params, timeout=20)
            if resp.status_code != 200:
                print(f"  ⚠️  Crossref returned HTTP {resp.status_code}, stopping.")
                break

            data = resp.json()
            items = data.get("message", {}).get("items", [])
            if not items:
                print("No more items returned.")
                break

            for item in items:
                authors = item.get("author", [])
                if not authors:
                    continue

                # Get publication year
                pub_date = item.get("published-print") or item.get("published-online") or {}
                year_parts = pub_date.get("date-parts", [[None]])
                year = year_parts[0][0] if year_parts and year_parts[0] else None

                # Get journal name
                container = item.get("container-title", [])
                journal = container[0] if container else "N/A"

                papers.append({
                    "doi":     item.get("DOI", "N/A"),
                    "title":   (item.get("title") or ["N/A"])[0],
                    "authors": authors,
                    "year":    year,
                    "journal": journal,
                })

            offset += rows_per_page
            print(f"  Fetched {len(papers)} papers so far...")

            if len(items) < rows_per_page:
                break  

        except requests.exceptions.Timeout:
            print("  ⚠️  Request timed out. Stopping fetch.")
            break
        except Exception as e:
            print(f"  ⚠️  Error: {e}")
            break

    print(f"✅ Total papers fetched: {len(papers)}")
    return papers[:max_papers]




def normalize_family_name(name: str) -> str:
    """
    Normalise an author family name for comparison:
      - Convert to lowercase
      - Remove punctuation and special characters
      - Strip whitespace
    """
    name = name.lower()
    name = re.sub(r"[^a-z\s]", "", name)
    return name.strip()


def is_alphabetical(authors: list) -> bool:
    """
    Return True if the author list is in ascending alphabetical order
    by family name (after normalisation).

    Returns False if any author is missing a family name.
    """
    family_names = []
    for author in authors:
        family = author.get("family", "").strip()
        if not family:
            return False  
        family_names.append(normalize_family_name(family))

    return family_names == sorted(family_names)


def chance_probability(n: int) -> float:
    """
    Probability that a randomly ordered list of n authors
    happens to be alphabetical = 1 / n!

    This quantifies the false-positive risk per paper.
    """
    return 1.0 / math.factorial(n)




def analyze_papers(papers: list, min_authors: int = 4):
    """
    Compute per-paper alphabetical flags and aggregate journal-level statistics.

    Parameters
    ----------
    papers      : list — output of fetch_papers_crossref()
    min_authors : int  — minimum author count to include a paper (default: 4)

    Returns
    -------
    paper_results : list of per-paper dicts
    summary       : dict with journal-level statistics and classification
    """
    paper_results = []
    alpha_count = 0
    total_eligible = 0
    author_contributions_hints = 0

    for paper in papers:
        authors = paper["authors"]
        n = len(authors)

        
        if n < min_authors:
            continue

        total_eligible += 1
        alpha = is_alphabetical(authors)
        if alpha:
            alpha_count += 1

        title_lower = paper["title"].lower()
        contrib_keywords = ["conceptualization", "methodology", "supervision",
                            "writing", "original draft", "review & editing"]
        has_contrib_hint = any(kw in title_lower for kw in contrib_keywords)
        if has_contrib_hint:
            author_contributions_hints += 1

        paper_results.append({
            "doi":             paper["doi"],
            "title":           paper["title"],
            "journal":         paper["journal"],
            "year":            paper["year"],
            "num_authors":     n,
            "author_names":    " | ".join(a.get("family", "?") for a in authors),
            "is_alphabetical": alpha,
            "chance_prob":     round(chance_probability(n), 8),
        })

   
    alpha_rate = alpha_count / total_eligible if total_eligible > 0 else 0.0


    if total_eligible < 20:
        conclusion = "Insufficient data"
        confidence = "Low"
    elif alpha_rate >= 0.75:
        conclusion = "Alphabetical-dominant"
        confidence = "High" if total_eligible >= 100 else "Medium"
    elif alpha_rate <= 0.25:
        conclusion = "Contribution-based"
        confidence = "High" if total_eligible >= 100 else "Medium"
    else:
        conclusion = "Mixed / Unclear"
        confidence = "Low"

    summary = {
        "total_papers_fetched":      len(papers),
        "eligible_papers":           total_eligible,
        "alphabetical_papers":       alpha_count,
        "contribution_based_papers": total_eligible - alpha_count,
        "alphabetical_rate":         round(alpha_rate, 4),
        "conclusion":                conclusion,
        "confidence":                confidence,
        "min_authors_filter":        min_authors,
    }

    return paper_results, summary




def save_fetched_data(papers: list, journal_label: str):
    """
    Save ALL fetched paper data (including papers with < min authors)
    to a styled Excel file — one row per paper.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Fetched Papers"

    dark_blue = "1F4E79"
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_align = Alignment(vertical="top", wrap_text=True)

    headers = [
        "No.", "DOI", "Title", "Journal", "Year",
        "# Authors", "Author Names (Family, Given)"
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    for i, paper in enumerate(papers, 1):
        authors = paper["authors"]
        author_str = " | ".join(
            f"{a.get('family', '?')}, {a.get('given', '?')}" for a in authors
        )
        ws.append([
            i,
            paper["doi"],
            paper["title"],
            paper["journal"],
            paper["year"],
            len(authors),
            author_str,
        ])
        for col in range(1, len(headers) + 1):
            ws.cell(ws.max_row, col).alignment = top_align

    # Column widths
    col_widths = [6, 35, 65, 35, 8, 10, 80]
    for col_idx, width in enumerate(col_widths, 1):
        letter = ws.cell(1, col_idx).column_letter
        ws.column_dimensions[letter].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = re.sub(r"[^\w]", "_", journal_label)[:80]
    filename = f"fetched_papers_{safe_label}_{timestamp}.xlsx"
    wb.save(filename)
    print(f"📁 Fetched paper data saved to: {filename}")
    return filename


def save_to_excel(paper_results: list, summary: dict, journal_label: str):
    """
    Save analysis results to a styled Excel file with two sheets:
      Sheet 1 — Journal Summary (classification result)
      Sheet 2 — Per-Paper Results (one row per paper)
    """
    wb = Workbook()

    
    dark_blue  = "1F4E79"
    mid_blue   = "366092"
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_align    = Alignment(vertical="top", wrap_text=True)

    def make_header_style(ws, row_num, bg_color):
        fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=12)
        for cell in ws[row_num]:
            cell.fill = fill
            cell.font = font
            cell.alignment = center_align


    ws1 = wb.active
    ws1.title = "Journal Summary"

    ws1.append(["Metric", "Value"])
    make_header_style(ws1, 1, dark_blue)

    conclusion = summary["conclusion"]
    if conclusion == "Contribution-based":
        result_fill = green_fill
    elif conclusion == "Alphabetical-dominant":
        result_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    else:
        result_fill = yellow_fill

    summary_rows = [
        ("Journal / Query",                   journal_label),
        ("Min Authors Filter",                f"≥ {summary['min_authors_filter']} authors"),
        ("Total Papers Fetched",              summary["total_papers_fetched"]),
        ("Eligible Papers (≥ min authors)",   summary["eligible_papers"]),
        ("Alphabetical Papers",               summary["alphabetical_papers"]),
        ("Contribution-based Papers",         summary["contribution_based_papers"]),
        ("AlphabeticalRate",                  f"{summary['alphabetical_rate']:.2%}"),
        ("Conclusion",                        conclusion),
        ("Confidence",                        summary["confidence"]),
    ]

    for metric, value in summary_rows:
        ws1.append([metric, value])
        row_idx = ws1.max_row
        if metric in ("Conclusion", "Confidence", "AlphabeticalRate"):
            for col in [1, 2]:
                ws1.cell(row_idx, col).fill = result_fill
                ws1.cell(row_idx, col).font = Font(bold=True, size=11)
        for col in [1, 2]:
            ws1.cell(row_idx, col).alignment = center_align

    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 35

  
    ws1.append([]) 
    ws1.append(["AlphabeticalRate", "Classification", "Confidence Level"])
    make_header_style(ws1, ws1.max_row, mid_blue)

    legend = [
        ("≥ 75%",       "Alphabetical-dominant", "High (if ≥ 100 papers)"),
        ("26% – 74%",   "Mixed / Unclear",        "Low"),
        ("≤ 25%",       "Contribution-based",     "High (if ≥ 100 papers)"),
    ]
    for row_data in legend:
        ws1.append(list(row_data))
        row_idx = ws1.max_row
        for col in range(1, 4):
            ws1.cell(row_idx, col).alignment = center_align


    ws2 = wb.create_sheet("Per-Paper Results")

    paper_headers = [
        "No.",
        "DOI",
        "Publication Title",
        "Journal",
        "Year",
        "# Authors",
        "Author Family Names (in order)",
        "Is Alphabetical?",
        "Chance Prob (1/n!)",
    ]
    ws2.append(paper_headers)
    make_header_style(ws2, 1, mid_blue)

    for i, p in enumerate(paper_results, 1):
        row_data = [
            i,
            p["doi"],
            p["title"],
            p["journal"],
            p["year"],
            p["num_authors"],
            p["author_names"],
            "Yes" if p["is_alphabetical"] else "No",
            p["chance_prob"],
        ]
        ws2.append(row_data)
        row_idx = ws2.max_row

        
        cell = ws2.cell(row_idx, 8)
        cell.fill = green_fill if p["is_alphabetical"] else red_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align

        for col in range(1, len(paper_headers) + 1):
            ws2.cell(row_idx, col).alignment = top_align

  
    col_widths = [5, 30, 60, 30, 8, 10, 65, 16, 18]
    for col_idx, width in enumerate(col_widths, 1):
        letter = ws2.cell(1, col_idx).column_letter
        ws2.column_dimensions[letter].width = width

   
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_label = re.sub(r"[^\w]", "_", journal_label)
    
    safe_label = safe_label[:80]
    filename = f"author_order_{safe_label}_{timestamp}.xlsx"
    wb.save(filename)

    print(f"\n📁 Results saved to: {filename}")
    return filename




def main():
    parser = argparse.ArgumentParser(
        description=(
            "Classify journal author ordering (alphabetical vs contribution-based) "
            "by analysing papers via the Crossref API."
        )
    )
    parser.add_argument(
        "--issn",
        type=str,
        help="Journal ISSN (e.g., 1476-4687 for Nature)"
    )
    parser.add_argument(
        "--journal",
        type=str,
        help="Journal name keyword (e.g., 'IEEE Transactions on Networking')"
    )
    parser.add_argument(
        "--max",
        type=int,
           default=1000,
           help="Maximum papers to fetch (default: 1000)"
    )
    parser.add_argument(
        "--min-authors",
        type=int,
        default=4,
        help="Minimum authors per paper to include (default: 4)"
    )
    args = parser.parse_args()

    if not args.issn and not args.journal:
        user_input = input("Enter journal name or ISSN: ").strip()
        if user_input.replace("-", "").isdigit() or (
            len(user_input) == 9 and user_input[4] == "-"
        ):
            args.issn = user_input
        else:
            args.journal = user_input

    journal_label = args.issn or args.journal

    print(f"\n{'='*60}")
    print(f"  Author Order Classifier")
    print(f"  Journal  : {journal_label}")
    print(f"  Max papers     : {args.max}")
    print(f"  Min authors    : {args.min_authors}+")
    print(f"{'='*60}")

    
    papers = fetch_papers_crossref(
        issn=args.issn,
        journal_name=args.journal,
        max_papers=args.max,
    )

    if not papers:
        print("❌ No papers found. Check the ISSN or journal name.")
        return

    # Save raw fetched data to Excel
    save_fetched_data(papers, journal_label)

   
    paper_results, summary = analyze_papers(papers, min_authors=args.min_authors)

   
    print(f"\n{'='*60}")
    print(f"  RESULTS")
    print(f"{'='*60}")
    print(f"  Total papers fetched        : {summary['total_papers_fetched']}")
    print(f"  Eligible papers (≥{args.min_authors} authors) : {summary['eligible_papers']}")
    print(f"  Alphabetical papers         : {summary['alphabetical_papers']}")
    print(f"  Contribution-based papers   : {summary['contribution_based_papers']}")
    print(f"  AlphabeticalRate            : {summary['alphabetical_rate']:.2%}")
    print(f"  Conclusion                  : {summary['conclusion']}")
    print(f"  Confidence                  : {summary['confidence']}")
    print(f"{'='*60}")

   
    save_to_excel(paper_results, summary, journal_label)


if __name__ == "__main__":
    main()
